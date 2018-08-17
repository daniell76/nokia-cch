import re
import openpyxl
import openpyxl.styles
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import *
from openpyxl.compat import range

VNF_COLUMN_NAMES = ('vnf', 'vnfNum', 'vdu', 'vduNum', 'vcpu', 'memory', 'storage', 'vmRed', 'hwAntiAffin')

def isInRange(irange, cell):
  ret = False
  min_col, min_row, max_col, max_row = range_boundaries(irange)
  if cell.row in range(min_row,max_row+1) and column_index_from_string(cell.column) in range(min_col,max_col+1):
    ret = True
  return ret

def isMergedCell(cell):
  ret = False
  if cell is None:
    return ret
  if cell.coordinate in cell.parent.merged_cells:
    ret = True
  return ret

def getMergedRange(cell):
  ret = None
  if cell == None:
    return ret
  if isMergedCell(cell):
    for irange in cell.parent.merged_cell_ranges:
      if isInRange(irange, cell):
        ret = irange
        break
  else:
    ret = cell.coordinate+':'+cell.coordinate
  return ret

def getRowRange(cell):
  irange = getMergedRange(cell)
  s,f = irange.split(':')
  sCol, sRow = coordinate_from_string(s)
  fCol, fRow = coordinate_from_string(f)
  return sRow,fRow

def getRowSpan(cell):
  sRow,fRow = getRowRange(cell)
  return fRow - sRow + 1

def getColumnRange(cell):
  irange = getMergedRange(cell)
  s,f = irange.split(':')
  sCol, sRow = coordinate_from_string(s)
  fCol, fRow = coordinate_from_string(f)
  return sCol, fCol

def getColumnSpan(cell):
  sCol, fCol = getColumnRange(cell)
  sColNum = column_index_from_string(sCol)
  fColNum = column_index_from_string(fCol)
  return fColNum - sColNum + 1

def getMergedCellValue(sheet, irange, connector=''):
  ret = []
  min_col, min_row, max_col, max_row = range_boundaries(irange)
  for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
      v = sheet.cell(row=row, column=col).value
      if v is not None:
        v = str(v).strip()
        if len(v) == 0 :continue
        ret.append(str(v))
  return connector.join(ret)

def getCellValueWithMergeLookup(cell, connector=''):
  if isMergedCell(cell):
    sheet = cell.parent
    return getMergedCellValue(sheet, getMergedRange(cell), connector)
  else:
    return cell.value

# Force to convert to an int  
def forceToInt(s):
  ret = 0
  try:
    ret = int(s)
  except:
    ret = 0
  return ret

# get the list of blocks
def getDataRanges(sheet, matchRe):
  ret = []
  for r in range(1,4):
    for cell in sheet[r]:
      if cell.value is None: continue
      p = re.compile(matchRe)
      m = p.search(str(cell.value))
      if not m: continue
      titleRange = getMergedRange(sheet, cell)
      min_col, min_row, max_col, max_row = range_boundaries(titleRange)
      min_row += 2
      max_row = sheet.max_row
      dataRange =  get_column_letter(min_col)+str(min_row)+':'+get_column_letter(max_col)+str(max_row)
      ret.append(dataRange)
  return ret

def setCellBgColor(cell, colorCode):
  cell.fill = openpyxl.styles.PatternFill(start_color=colorCode, end_color=colorCode, fill_type = "solid")

def setCellFgColor(cell, colorCode):
  cell.font = openpyxl.styles.Font(name='Calibri', color=colorCode, bold=True)

def setCellAlignment(cell, horizontal=None, vertical=None):
  cell.alignment=openpyxl.styles.Alignment(horizontal=horizontal, vertical=vertical)
  
  